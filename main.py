from openpyxl import Workbook
from openpyxl import load_workbook
import difflib
import heapq

if __name__ == "__main__":
    # read from the terminal:
    command = (list(input("Please enter k and the string you want to compare").split()))
    k = command[0]  # the first is topk
    lens = len(command)
    print("lens =", lens)

    # getExcel:
    # 实例化
    wb = Workbook()
    # 激活 worksheet
    ws = wb.active

    wb2 = load_workbook('测试.xlsx')
    print(wb2.sheetnames)

    sheet = wb2["关联关系"]
    max_row = sheet.max_row
    max_col = sheet.max_column

    # creat Array:
    board = ['s' for i in range(max_row)]
    similarities_of_data_element = [0 for i in range(max_row)]

    # read the data element:
    count = 0
    for row in sheet.rows:
        board[count] = row[3].value
        count += 1

    # only have : topk \ B \ D
    if lens == 3:
        for i in range(0, max_row):
            similarities_of_data_element[i] = difflib.SequenceMatcher(None, command[2], board[i]).ratio()
    # only have : topk \ A \ B \ D
    elif lens == 4:
        for i in range(0, max_row):
            similarities_of_data_element[i] = difflib.SequenceMatcher(None, command[3], board[i]).ratio()

    # choose the top10k:
    num_of_chosen = 10 * int(k)
    if num_of_chosen > max_row:
        num_of_chosen = max_row
    num_dict = {}
    for i in range(len(similarities_of_data_element)):
        num_dict[i] = similarities_of_data_element[i]
    res_list = sorted(num_dict.items(), key=lambda e: e[1])
    data_element_largestTopk_index = list([one[0] for one in res_list[::-1][:num_of_chosen]])

    ###

    for i in range(0, num_of_chosen):
        print(data_element_largestTopk_index[i], end=' ')
        print(similarities_of_data_element[data_element_largestTopk_index[i]])
    ###

    # compare the largest similarities' topk:
    similarities_of_institute_name = ['s' for i in range(0, num_of_chosen)]
    similarities_of_table_Chinese_name = ['s' for i in range(0, num_of_chosen)]

    # read the institute name/A:
    count = 0
    for row in sheet.rows:
        if row[0].value != None:
            board[count] = row[0].value
            print(board[i])
        count += 1
    # only have : topk \ B \ D
    if lens == 3:
        for i in range(0, num_of_chosen):
            similarities_of_institute_name[i] = 0
    # only have : topk \ A \ B \ D
    elif lens == 4:
        for i in range(0, num_of_chosen):
            similarities_of_institute_name[i] = difflib.SequenceMatcher(None, command[1], board[data_element_largestTopk_index[i]]).ratio()