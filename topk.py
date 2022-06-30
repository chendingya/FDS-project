from openpyxl import Workbook
from openpyxl import load_workbook
import difflib
import time
import re
from simhash import Simhash


def topk(todo):
    # start
    timeStart = time.time()
    command = list(todo.split())
    k = command[0]  # the first is topk
    lens = len(command)

    # getExcel:
    # 实例化
    wb = Workbook()
    # 激活 worksheet
    ws = wb.active

    wb2 = load_workbook('测试.xlsx')

    sheet = wb2["关联关系"]
    max_row = sheet.max_row

    # creat Array:
    boardD = ['s' for i in range(max_row)]
    boardA = ['s' for i in range(max_row)]
    boardB = ['s' for i in range(max_row)]
    similarities_of_data_element = [0 for i in range(max_row)]

    # read the data element:
    count = 0
    for row in sheet.rows:
        boardD[count] = row[3].value
        count += 1

    # only have : topk \ B \ D
    if lens == 3:
        for i in range(0, max_row):
            similarities_of_data_element[i] = difflib.SequenceMatcher(None, command[2], boardD[i]).ratio()
    # only have : topk \ A \ B \ D
    if lens == 4:
        for i in range(0, max_row):
            similarities_of_data_element[i] = difflib.SequenceMatcher(None, command[3], boardD[i]).ratio()

    # only have : topk \ B \ D
    if lens == 3:
        for i in range(0, max_row):
            a_simhash = Simhash(command[2])
            b_simhash = Simhash(boardD[i])
            max_hashbit = max(len(bin(a_simhash.value)), len(bin(b_simhash.value)))
            # 汉明距离
            distince = a_simhash.distance(b_simhash)
            similar = 1 - distince / max_hashbit
            similarities_of_data_element[i] = (similar + similarities_of_data_element[i]) / 2
    # only have : topk \ A \ B \ D
    if lens == 4:
        for i in range(0, max_row):
            a_simhash = Simhash(command[3])
            b_simhash = Simhash(boardD[i])
            max_hashbit = max(len(bin(a_simhash.value)), len(bin(b_simhash.value)))
            # 汉明距离
            distince = a_simhash.distance(b_simhash)
            similar = 1 - distince / max_hashbit
            similarities_of_data_element[i] = (similar + similarities_of_data_element[i]) / 2

    # choose the top50percent:
    num_of_chosen = 0
    numk = int(k)
    for i in range(0, max_row):
        if similarities_of_data_element[i] >= 0.5:
            num_of_chosen += 1

    if numk > num_of_chosen:
        num_of_chosen = numk
    num_dict = {}
    for i in range(len(similarities_of_data_element)):
        num_dict[i] = similarities_of_data_element[i]
    res_list = sorted(num_dict.items(), key=lambda e: e[1])
    while similarities_of_data_element[max_row - num_of_chosen - 1] == similarities_of_data_element[
        max_row - num_of_chosen] and max_row - num_of_chosen - 1 >= 0:
        num_of_chosen = num_of_chosen + 1
    data_element_largestTopk_index = list([one[0] for one in res_list[::-1][:num_of_chosen]])

    # compare the largest similarities' topk:
    similarities_of_institute_name = [0 for i in range(0, num_of_chosen)]
    similarities_of_table_Chinese_name = [0 for i in range(0, num_of_chosen)]

    # read the institute name/A:
    count = 0
    for i in range(0, max_row):
        boardA[i] = None
    for row in sheet.rows:
        if row[0].value is not None:
            boardA[count] = row[0].value
        count += 1
    # only have : topk \ B \ D
    if lens == 3:
        for i in range(0, num_of_chosen):
            if boardA[data_element_largestTopk_index[i]] is not None:
                similarities_of_institute_name[i] = 0
            else:
                similarities_of_institute_name[i] = 1
    # only have : topk \ A \ B \ D
    elif lens == 4:
        for i in range(0, num_of_chosen):
            if boardA[data_element_largestTopk_index[i]] is not None:
                similarities_of_institute_name[i] = difflib.SequenceMatcher(None, command[1], boardA[
                    data_element_largestTopk_index[i]]).ratio()
                a_simhash = Simhash(command[1])
                b_simhash = Simhash(boardA[data_element_largestTopk_index[i]])
                max_hashbit = max(len(bin(a_simhash.value)), len(bin(b_simhash.value)))
                # 汉明距离
                distince = a_simhash.distance(b_simhash)
                similar = 1 - distince / max_hashbit

                similarities_of_institute_name[i] = (similarities_of_institute_name[i] + similar ) / 2

            else:
                similarities_of_institute_name[i] = 0
    # read the table Chinese name/B:
    count = 0
    for i in range(0, max_row):
        boardB[i] = None
    for row in sheet.rows:
        boardB[count] = row[1].value
        count += 1
    for i in range(0, num_of_chosen):
        similarities_of_table_Chinese_name[i] = difflib.SequenceMatcher(None, command[2], boardB[
            data_element_largestTopk_index[i]]).ratio()
        a_simhash = Simhash(command[2])
        b_simhash = Simhash(boardB[data_element_largestTopk_index[i]])
        max_hashbit = max(len(bin(a_simhash.value)), len(bin(b_simhash.value)))
        # 汉明距离
        distince = a_simhash.distance(b_simhash)
        similar = 1 - distince / max_hashbit
        similarities_of_table_Chinese_name[i] = (similarities_of_table_Chinese_name[i] + similar) / 2

    # calculate the similarities:
    similarities = [0 for i in range(0, num_of_chosen)]
    for i in range(0, num_of_chosen):
        similarities[i] = (similarities_of_table_Chinese_name[i] + similarities_of_institute_name[i] +
                           similarities_of_data_element[data_element_largestTopk_index[i]]) / 3

    ###

    # rank the topk:
    num_of_topk = int(k)
    num_sort = {}
    for i in range(len(similarities)):
        num_sort[i] = similarities[i]
    res_list2 = sorted(num_sort.items(), key=lambda e: e[1])
    largestTopk_index = list([one[0] for one in res_list2[::-1][:num_of_topk]])

    # print:
    for i in range(0, num_of_topk):
        print(data_element_largestTopk_index[largestTopk_index[i]] + 1,
              boardA[data_element_largestTopk_index[largestTopk_index[i]]],
              boardB[data_element_largestTopk_index[largestTopk_index[i]]],
              boardD[data_element_largestTopk_index[largestTopk_index[i]]], similarities[largestTopk_index[i]])

    # end
    timeEnd = time.time()
    runTime = timeEnd - timeStart
    print("The running time is " + str(runTime) + "s", end="")
    result = str(data_element_largestTopk_index[largestTopk_index[0]]) + " " + str(runTime)
    for i in range(0, num_of_topk):
        result = result + " " + str(data_element_largestTopk_index[largestTopk_index[i]])
    return result


if __name__ == "__main__":
    # read from the terminal:
    string = input("Please enter k and the string you want to compare\n")

    topk(string)
