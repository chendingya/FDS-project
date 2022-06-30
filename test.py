import topk
from openpyxl import Workbook
from openpyxl import load_workbook

# getExcel:
# 实例化
wb = Workbook()
# 激活 worksheet
ws = wb.active

wb2 = load_workbook('测试.xlsx')

sheet = wb2["关联关系"]
max_row = sheet.max_row
max_col = sheet.max_column

# creat Array:
boardC = ['s' for i in range(max_row)]
boardA = ['s' for i in range(max_row)]
boardB = ['s' for i in range(max_row)]

# read the data element:
count = 0
for row in sheet.rows:
    boardC[count] = row[2].value
    count += 1

# read the institute name/A:
count = 0
for i in range(0, max_row):
    boardA[i] = None
for row in sheet.rows:
    if row[0].value is not None:
        boardA[count] = row[0].value
    count += 1

# read the table Chinese name/B:
count = 0
for i in range(0, max_row):
    boardB[i] = None
for row in sheet.rows:
    boardB[count] = row[1].value
    count += 1

number = int(input("Please enter k\n"))

count_of_true = 0
sum_runtime = 0
count_of_true_topk = 0
for i in range(0, max_row):
    if boardA[i] is not None:
        result = topk.topk(str(number) + ' ' + boardA[i] + " " + boardB[i] + " " + boardC[i]).split()
    else:
        result = topk.topk(str(number) + " " + boardB[i] + " " + boardC[i]).split()
    index = result[0]
    runtime = result[1]
    if int(index) == i:
        count_of_true = count_of_true + 1
        print(True)
    else:
        print(False)
    print()
    sum_runtime = sum_runtime + float(runtime)
    for j in range(2, len(result)):
        if int(result[j]) == i:
            count_of_true_topk = count_of_true_topk + 1
            break
accuracy = count_of_true / max_row
topk_accuracy = count_of_true_topk / max_row
average_runtime = sum_runtime / max_row
print("The accuracy is " + str(accuracy) + ".\n" + "The topk_accuracy is " + str(topk_accuracy) + ".\n" +
      "The average runtime is " + str(average_runtime) + 's.', end="")
